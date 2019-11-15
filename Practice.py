# import openpyxl and tkinter modules 
from openpyxl import *
from tkinter import *

# globally declare wb and sheet variable 

# opening the existing excel file 
wb = load_workbook('/Users/youshamahamuni2/Desktop/PythonClass1/Project1.xlsx') 

# create the sheet object 
sheet = wb.active 


def excel(): 
	
	# resize the width of columns in 
	# excel spreadsheet 
	sheet.column_dimensions['A'].width = 30
	sheet.column_dimensions['B'].width = 10
	sheet.column_dimensions['C'].width = 10
	sheet.column_dimensions['D'].width = 20


	# write given data to an excel spreadsheet 
	# at particular location 
	sheet.cell(row=1, column=1).value = "Name of Event"
	sheet.cell(row=1, column=2).value = "Date of Event"
	sheet.cell(row=1, column=3).value = "Type of Event"
	sheet.cell(row=1, column=4).value = "No.of People"



# Function to set focus (cursor) 
def focus1(event): 
	# set focus on the date1_field box 
	date1_field.focus_set() 


# Function to set focus 
def focus2(event): 
	# set focus on the sem_field box 
	sem_field.focus_set() 


# Function to set focus 
def focus3(event): 
	# set focus on the form_no_field box 
	form_no_field.focus_set() 


# Function to set focus 
def focus4(event): 
	# set focus on the contact_no_field box 
	contact_no_field.focus_set() 





# Function for clearing the 
# contents of text entry boxes 
def clear(): 
	
	# clear the content of text entry box 
	name_field.delete(0, END) 
	date1_field.delete(0, END) 
	sem_field.delete(0, END) 
	No_ppl_field.delete(0, END) 
	


# Function to take data from GUI 
# window and write to an excel file 
def insert(): 
	
	# if user not fill any entry 
	# then print "empty input" 
	if (name_field.get() == "" and
		date1_field.get() == "" and
		sem_field.get() == "" and
		form_no_field.get() == ""): 
			
		print("empty input") 

	else: 

		# assigning the max row and max column 
		# value upto which data is written 
		# in an excel sheet to the variable 
		current_row = sheet.max_row 
		current_column = sheet.max_column 

		# get method returns current text 
		# as string which we write into 
		# excel spreadsheet at particular location 
		sheet.cell(row=current_row + 1, column=1).value = name_field.get() 
		sheet.cell(row=current_row + 1, column=2).value = date1_field.get() 
		sheet.cell(row=current_row + 1, column=3).value = sem_field.get() 
		sheet.cell(row=current_row + 1, column=4).value = No_ppl_field.get() 
		

		# save the file 
		wb.save('/Users/youshamahamuni2/Desktop/PythonClass1/Project1.xlsx')

		# set focus on the name_field box 
		name_field.focus_set() 

		# call the clear() function 
		clear() 


# Driver code 
if __name__ == "__main__": 
	
	# create a GUI window 
	root = Tk() 

	# set the background colour of GUI window 
	root.configure(background='#7CBBD9') 

	# set the title of GUI window 
	root.title("Event Management Form") 

	# set the configuration of GUI window 
	root.geometry("600x300") 

	excel() 

	# create a Form label 
	heading = Label(root, text="Event Management Form", bg="#7CBBD9") 

	# create a Name label 
	name = Label(root, text="Name of Event", bg="#7CBBD9") 

	# create a date1 label 
	date1 = Label(root, text="Date of Event(DD/MM/YY)", bg="#7CBBD9") 

	# create a Semester label 
	sem = Label(root, text="Type of Event", bg="#7CBBD9") 

	# create a Form No. lable 
	no_ppl = Label(root, text="No. of People", bg="#7CBBD9") 

	
	# create a address label 
	

	# grid method is used for placing 
	# the widgets at respective positions 
	# in table like structure . 
	heading.grid(row=0, column=1) 
	name.grid(row=1, column=0) 
	date1.grid(row=2, column=0) 
	sem.grid(row=3, column=0) 
	no_ppl.grid(row=4,column=0)
	

	# create a text entry box 
	# for typing the information 
	name_field = Entry(root) 
	date1_field = Entry(root) 
	sem_field = Entry(root) 
	No_ppl_field = Entry(root) 


	# bind method of widget is used for 
	# the binding the function with the events 

	# whenever the enter key is pressed 
	# then call the focus1 function 
	name_field.bind("<Return>", focus1) 

	# whenever the enter key is pressed 
	# then call the focus2 function 
	date1_field.bind("<Return>", focus2) 

	# whenever the enter key is pressed 
	# then call the focus3 function 
	sem_field.bind("<Return>", focus3) 

	# whenever the enter key is pressed 
	# then call the focus4 function 
	No_ppl_field.bind("<Return>", focus4) 

	# whenever the enter key is pressed 
	# then call the focus5 function 


	# whenever the enter key is pressed 
	# then call the focus6 function 
	 

	# grid method is used for placing 
	# the widgets at respective positions 
	# in table like structure . 
	name_field.grid(row=1, column=1, ipadx="100") 
	date1_field.grid(row=2, column=1, ipadx="100") 
	sem_field.grid(row=3, column=1, ipadx="100") 
	No_ppl_field.grid(row=4, column=1, ipadx="100") 
	

	# call excel function 
	excel() 

	# create a Submit Button and place into the root window 
	submit = Button(root, text="Submit", fg="Red", 
							bg="Black", command=insert) 
	submit.grid(row=8, column=1) 

	# start the GUI 
	root.mainloop() 
